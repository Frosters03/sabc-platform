from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, timedelta
import io

from core.database import get_db
from models.pointage import Pointage, LignePointage
from models.equipe import Equipe
from models.utilisateur import Utilisateur
from schemas.pointage import PointageCreate, PointageResponse
from .auth import get_current_user, require_role, log_action

router = APIRouter(prefix="/pointages", tags=["Pointages"])

HEURES_QUART = {
    "7h-19h":  12,
    "19h-7h":  12,
    "7h-14h":  7,
    "14h-21h": 7,
    "21h-7h":  10,
}


@router.post("/", response_model=PointageResponse)
def creer_pointage(
    data: PointageCreate,
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    # Vérifier double saisie
    existant = db.query(Pointage).filter(
        Pointage.date      == data.date,
        Pointage.equipe_id == data.equipe_id,
        Pointage.quart     == data.quart,
    ).first()
    if existant:
        raise HTTPException(
            status_code=400,
            detail=f"Pointage déjà saisi pour cette équipe le {data.date} — quart {data.quart}"
        )

    pointage = Pointage(
        date       = data.date,
        chaine     = data.chaine,
        equipe_id  = data.equipe_id,
        equipe_nom = data.equipe_nom,
        quart      = data.quart,
        saisi_par  = current_user.username,
    )
    db.add(pointage)
    db.flush()

    heures_quart = HEURES_QUART.get(data.quart, 8)

    for ligne in data.lignes:
        heures_N = heures_quart if ligne.presence == "P" else 0
        ligne_db = LignePointage(
            pointage_id     = pointage.id,
            membre_id       = ligne.membre_id,
            fonction        = ligne.fonction,
            nom_prenom      = ligne.nom_prenom,
            statut_emploi   = ligne.statut_emploi,
            presence        = ligne.presence,
            heures_N        = heures_N,
            heures_F        = ligne.heures_F,
            heures_PN       = ligne.heures_PN,
            est_occasionnel = ligne.est_occasionnel,
        )
        db.add(ligne_db)

    db.commit()
    db.refresh(pointage)
    log_action(db, current_user.username, "INSERT", "pointages",
               f"Pointage {data.date} — {data.equipe_nom} — {data.quart}")
    return pointage


@router.get("/", response_model=List[PointageResponse])
def get_pointages(
    equipe_id:  Optional[int]  = Query(None),
    chaine:     Optional[str]  = Query(None),
    date_debut: Optional[date] = Query(None),
    date_fin:   Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    query = db.query(Pointage)
    if equipe_id:  query = query.filter(Pointage.equipe_id == equipe_id)
    if chaine:     query = query.filter(Pointage.chaine    == chaine)
    if date_debut: query = query.filter(Pointage.date      >= date_debut)
    if date_fin:   query = query.filter(Pointage.date      <= date_fin)
    return query.order_by(Pointage.date.desc()).all()


@router.get("/verifier")
def verifier_pointage(
    date:       date = Query(...),
    equipe_id:  int  = Query(...),
    quart:      str  = Query(...),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    existant = db.query(Pointage).filter(
        Pointage.date      == date,
        Pointage.equipe_id == equipe_id,
        Pointage.quart     == quart,
    ).first()
    return {"existe": existant is not None, "id": existant.id if existant else None}


@router.get("/export-excel")
def export_excel_semaine(
    equipe_id:  int  = Query(...),
    date_lundi: date = Query(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta
    import io

    date_dimanche = date_lundi + timedelta(days=6)
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Équipe non trouvée")

    pointages = db.query(Pointage).filter(
        Pointage.equipe_id == equipe_id,
        Pointage.date      >= date_lundi,
        Pointage.date      <= date_dimanche,
    ).all()

    par_date = {}
    for p in pointages:
        d = p.date
        par_date[d] = p

    membres = sorted(equipe.membres, key=lambda m: m.ordre)

    wb = openpyxl.Workbook()
    ws = wb.active
    sem_num = date_lundi.isocalendar()[1]
    ws.title = f"SEM{sem_num:02d}"

    # ── STYLES ────────────────────────────────────────────
    thin  = Side(style="thin",   color="000000")
    thick = Side(style="medium", color="000000")
    b_all   = Border(top=thick,  bottom=thick,  left=thick,  right=thick)
    b_thick = Border(top=thick, bottom=thick, left=thick, right=thick)
    dashed = Side(style="dashed", color="000000")  # trait interrompu
    none  = Side(style=None)

    def st(cell, bold=False, size=10, color="000000", fill=None,
           halign="center", valign="center", border=None, wrap=False, italic=False):
        cell.font      = Font(bold=True, size=size, name="Arial",
                              color=color, italic=italic)
        cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                   wrap_text=wrap)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if border:
            cell.border = border

    def mc(r1, c1, r2, c2, value="", **kwargs):
        # 1. Fusionner les cellules
        ws.merge_cells(start_row=r1, start_column=c1,
                    end_row=r2,   end_column=c2)

        # 2. Écrire la valeur dans la première cellule
        cell_top = ws.cell(row=r1, column=c1, value=value)
        st(cell_top, **kwargs)   # style sur la première cellule

        # 3. Appliquer le style (surtout les bordures) à toutes les cellules de la plage
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                cell = ws.cell(row=row, column=col)
                # On applique le même style (sauf la valeur) pour assurer la continuité
                st(cell, **kwargs)

        return cell_top

    def wc(r, c, value="", **kwargs):
        """Write + style single cell"""
        cell = ws.cell(row=r, column=c, value=value)
        st(cell, **kwargs)
        return cell

    # ── LARGEURS COLONNES (fidèles au modèle) ─────────────
    ws.column_dimensions["A"].width  = 8.7
    ws.column_dimensions["B"].width  = 23.0
    for col_letter in ["C","D","E","F","G","H","I","J","K",
                       "L","M","N","O","P","Q","R","S","T","U","V"]:
        ws.column_dimensions[col_letter].width = 6.5
    ws.column_dimensions["W"].width  = 9.5
    ws.column_dimensions["X"].width  = 7.5
    ws.column_dimensions["Y"].width  = 7.5
    ws.column_dimensions["Z"].width  = 5.5
    ws.column_dimensions["AA"].width = 5.5
    ws.column_dimensions["AB"].width = 7.7

    # ── HAUTEURS LIGNES ───────────────────────────────────
    for r in range(1, 7):
        ws.row_dimensions[6].height = 0
    ws.row_dimensions[7].height  = 20.25
    ws.row_dimensions[8].height  = 15.75
    ws.row_dimensions[9].height  = 15.75
    ws.row_dimensions[10].height = 15.75
    for r in range(11, 32):
        ws.row_dimensions[r].height = 16.5
    ws.row_dimensions[32].height = 17.25
    ws.row_dimensions[33].height = 17.25
    ws.row_dimensions[34].height = 15.75
    for r in range(35, 40):
        ws.row_dimensions[r].height = 15.75
    ws.row_dimensions[39].height = 39.75

    # ── LIGNE 1-5 : En-têtes titre (comme sur la photo) ───
    # USINES DE DOUALA — col A-B lignes 1-4
    mc(1,1, 4,2, "USINES DE DOUALA",
       bold=True, size=10, halign="center", wrap=True, border=b_thick)

    # POINTAGE PERSONNEL — centré lignes 1-2
    mc(1,3, 2,28, "POINTAGE PERSONNEL",
       bold=True, size=20, halign="center", border=Border(top=thick, bottom=none, left=thick, right=thick))

    # Service + Atelier — ligne 3
    mc(3,3, 3,28,
       "Service : conditionnement                    Atelier : usine",
       size=10, halign="center", border=Border(top=none, bottom=none, left=thick, right=thick))

    # ── LIGNE 4 : NDOKOTI + Chaîne + Équipe ──────────────
    mc(4,3, 4,13,
       f"NDOKOTI {equipe.chaine}",
       bold=True, size=10, halign="center", border=Border(top=none, bottom=thick, left=thick, right=none))
    mc(4,14, 4,28,
       f"Equipe : {equipe.nom}",
       bold=True, size=10, halign="center", border=Border(top=none, bottom=thick, left=none, right=thick))

    # ── LIGNE 5 : SEM + Mois + Semaine du ────────────────
    mois_fr = ["","Janvier","Février","Mars","Avril","Mai","Juin",
               "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    wc(5,1, f"SEM:{sem_num:02d}", bold=True, size=9, halign="center")

    mc(5,3, 5,8,  "Mois",  size=9, halign="center")
    mc(5,9, 5,12,
       mois_fr[date_lundi.month],
       bold=True, size=9, halign="center")
    mc(5,13, 5,28,
       f"Semaine du : {date_lundi.strftime('%d/%m/%Y')}  {date_dimanche.strftime('%d/%m/%Y')}",
       bold=True, size=9, halign="center")

    # ── LIGNE 6 : séparateur vide ─────────────────────────
    mc(6,1, 6,28, "", border=none)

    # ── LIGNE 7-8-9 : En-têtes tableau ────────────────
    # MAT. — A7:A9
    mc(7,1, 9,1, "MAT.", bold=True, size=9,
       halign="center", border=b_thick)
    # NOMS ET PRENOMS — B7:B9
    mc(7,2, 9,2, "NOMS ET PRENOMS", bold=True, size=9,
       halign="left", border=b_thick, wrap=True)

    # Jours : C=3, F=6, I=9, L=12, O=15, R=18, U=21
    jours_cols = [
        ("lundi",     3,  5,  date_lundi + timedelta(days=0)),
        ("mardi",     6,  8,  date_lundi + timedelta(days=1)),
        ("mercredi",  9,  11, date_lundi + timedelta(days=2)),
        ("jeudi",     12, 14, date_lundi + timedelta(days=3)),
        ("vendredi",  15, 17, date_lundi + timedelta(days=4)),
        ("samedi",    18, 20, date_lundi + timedelta(days=5)),
        ("dimanche",  21, 22, date_lundi + timedelta(days=6)),
    ]

    for jour, c1, c2, d in jours_cols:
        # Ligne 7 : nom du jour fusionné
        mc(7, c1, 7, c2, jour.capitalize(),
           bold=True, size=9, halign="center", border=Border(top=thick, bottom=dashed, left=thick, right=thick))
        # Ligne 8 : date fusionnée
        mc(8, c1, 8, c2, d.strftime("%d/%m/%Y"),
           bold=True, size=8, halign="center", border=Border(top=dashed, bottom=dashed, left=thick, right=thick))
        # Ligne 9 : N/F/PN (dimanche = F/PN seulement)
        if jour == "dimanche":
            wc(9, c1, "F",  bold=True, size=9,
               halign="center", border=Border(top=dashed, bottom=thick, left=thick, right=dashed))
            wc(9, c2, "PN", bold=True, size=9,
               halign="center", border=Border(top=dashed, bottom=thick, left=dashed, right=thick))
        else:
            wc(9, c1,   "N",  bold=True, size=9, halign="center", border=Border(top=dashed, bottom=thick, left=thick, right=dashed))
            wc(9, c1+1, "F",  bold=True, size=9, halign="center", border=Border(top=dashed, bottom=thick, left=dashed, right=dashed))
            wc(9, c2,   "PN", bold=True, size=9, halign="center", border=Border(top=dashed, bottom=thick, left=dashed, right=thick))

    # REPARTITION — W7:AB7
    mc(7, 23, 7, 28, "REPARTITION",
       bold=True, size=9, halign="center", border=b_thick)

    # Sous-en-têtes répartition — lignes 8-9
    rep = [
        (23, "Heures\ntravaillées"),
        (24, "Heures\npayées"),
        (25, "Heures\nnorm"),
        (26, "HS"),
        (27, "PN"),
        (28, "Taux HS"),
    ]
    for col, lbl in rep:
        mc(8, col, 9, col, lbl,
           bold=True, size=8, halign="center",
           border=b_thick, wrap=True)

    # ── LIGNES MEMBRES 10-29 ──────────────────────────────
    ROW_S = 10
    ROW_E = 29

    for idx in range(20):
        r = ROW_S + idx
        # Appliquer borders sur toutes les cellules de la ligne
        for col in range(1, 29):
            ws.cell(row=r, column=col).border = b_all

        if idx < len(membres):
            m = membres[idx]

            # Récupérer présences par jour
            presences = {}
            for j_idx, (_, c1, c2, d) in enumerate(jours_cols):
                ptg = par_date.get(d.date() if hasattr(d,'date') else d)
                if ptg:
                    for lg in ptg.lignes:
                        if lg.membre_id == m.id and not lg.est_occasionnel:
                            presences[j_idx] = lg
                            break

            wc(r, 1, m.matricule or "", size=9, halign="center", border=Border(top=dashed, bottom=dashed, left=thick, right=thick))
            wc(r, 2, m.nom_prenom,      size=9, halign="left",   border=Border(top=dashed, bottom=dashed, left=thick, right=thick))

            for j_idx, (jour, c1, c2, d) in enumerate(jours_cols):
                lg = presences.get(j_idx)
                if jour == "dimanche":
                    f_val  = lg.heures_F  if lg and lg.heures_F  else ""
                    pn_val = lg.heures_PN if lg and lg.heures_PN else ""
                    wc(r, c1, f_val,  size=9, halign="center", border=Border(top=thick, bottom=thick, left=thick, right=dashed))
                    wc(r, c2, pn_val, size=9, halign="center", border=Border(top=thick, bottom=thick, left=dashed, right=thick))
                else:
                    if lg:
                        n_val  = lg.presence if lg.presence in ("AA","AB","R","RM","CP") \
                                 else (lg.heures_N if lg.heures_N else "")
                        f_val  = lg.heures_F  if lg.heures_F  else ""
                        pn_val = lg.heures_PN if lg.heures_PN else ""
                    else:
                        n_val = f_val = pn_val = ""
                    wc(r, c1,   n_val,  size=9, halign="center", border=Border(top=thick, bottom=thick, left=thick, right=dashed))
                    wc(r, c1+1, f_val,  size=9, halign="center", border=Border(top=thick, bottom=thick, left=dashed, right=dashed))
                    wc(r, c2,   pn_val, size=9, halign="center", border=Border(top=thick, bottom=thick, left=dashed, right=thick))
        else:
            wc(r, 1, "", border=b_all)
            wc(r, 2, "", border=b_all)
            for _, c1, c2, _ in jours_cols:
                for ci in range(c1, c2+1):
                    wc(r, ci, "", border=b_all)

        # Formules répartition
        rs = str(r)
        ws.cell(r,23).value = f"=SUM(C{rs}:D{rs},F{rs}:G{rs},I{rs}:J{rs},L{rs}:M{rs},O{rs}:P{rs},R{rs}:S{rs},U{rs})"
        ws.cell(r,24).value = f"=W{rs}+(0.5*(SUM(D{rs},G{rs},J{rs},M{rs},P{rs},S{rs},U{rs})))"
        ws.cell(r,25).value = f"=X{rs}-Z{rs}"
        ws.cell(r,26).value = f'=IF(X{rs}-40<0,"0",X{rs}-40)'
        ws.cell(r,27).value = f"=SUM(V{rs},E{rs},Q{rs},N{rs},K{rs},H{rs},T{rs})"
        ws.cell(r,28).value = f"=+Z{rs}/40"
        ws.cell(r,28).number_format = "0%"
        for ci in range(23, 29):
            st(ws.cell(r,ci), size=9, halign="center", border=b_all)

    # ── LIGNE 30 : TOTAUX ─────────────────────────────────
    R_T = 30
    mc(R_T, 1, R_T, 2, "", border=b_thick)
    for col in range(3, 29):
        cl = get_column_letter(col)
        ws.cell(R_T, col).value = f"=SUM({cl}{ROW_S}:{cl}{ROW_E})"
        st(ws.cell(R_T,col), bold=True, size=9,
           halign="center", border=b_thick)

    # ── LIGNES 31-32 : LÉGENDE ────────────────────────────
    mc(31,1,  31,2,  "RM=Maladie",              size=8, halign="left")
    mc(31,3,  31,8,  "CE= Congé événement familial", size=8, halign="left")
    mc(31,21, 31,22, "TAUX H.S",                bold=True, size=8, halign="center")
    mc(31,23, 31,28, f"=Z{R_T}/Y{R_T}",
       bold=True, size=9, halign="center")
    ws.cell(31,23).number_format = "0%"

    mc(32,1,  32,2,  "A=Absence",   size=8, halign="left")
    mc(32,3,  32,8,  "F=S150=Férié", size=8, halign="left")

    # ── LIGNES 33-38 : OBSERVATIONS + VISAS ──────────────
    mc(33,1,  38,16, "Observations", bold=True, size=9,
       halign="left", valign="top", border=b_thick)
    mc(33,17, 33,20, "Visa Contremaître",      size=8, halign="center", border=b_thick)
    mc(33,21, 33,23, "Visa Adj. Chef Service", size=8, halign="center", border=b_thick)
    mc(33,24, 33,28, "Visa Chef Service",      size=8, halign="center", border=b_thick)
    mc(34,17, 38,20, "", border=b_thick)
    mc(34,21, 38,23, "", border=b_thick)
    mc(34,24, 38,28, "", border=b_thick)

    # ── MISE EN PAGE ──────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A3
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = "A1:AB39"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (f"Pointage_{equipe.nom}_{equipe.chaine}_"
                f"SEM{sem_num:02d}_{date_lundi.strftime('%Y-%m-%d')}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export-excel-am")
def export_excel_am(
    equipe_id:  int  = Query(...),
    date_lundi: date = Query(...),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta
    import io

    date_dimanche = date_lundi + timedelta(days=6)
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Équipe non trouvée")

    pointages = db.query(Pointage).filter(
        Pointage.equipe_id == equipe_id,
        Pointage.date      >= date_lundi,
        Pointage.date      <= date_dimanche,
    ).all()
    par_date = {p.date: p for p in pointages}

    membres_am = sorted(
        [m for m in equipe.membres if m.statut == 'am' and m.actif != False],
        key=lambda m: m.ordre
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    sem_num = date_lundi.isocalendar()[1]
    ws.title = f"SEM{sem_num:02d}"

    thin   = Side(style="thin",   color="000000")
    thick  = Side(style="medium", color="000000")
    dashed = Side(style="dashed", color="000000")
    b_thick = Border(top=thick, bottom=thick, left=thick, right=thick)
    b_thin  = Border(top=thin,  bottom=thin,  left=thin,  right=thin)

    def st(cell, bold=False, size=10, color="000000",
           halign="center", valign="center", border=None, wrap=False):
        cell.font      = Font(bold=bold, size=size, name="Arial", color=color)
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if border: cell.border = border

    def mc(r1, c1, r2, c2, value="", **kwargs):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell_top = ws.cell(row=r1, column=c1, value=value)
        st(cell_top, **kwargs)
        for row in range(r1, r2+1):
            for col in range(c1, c2+1):
                st(ws.cell(row=row, column=col), **kwargs)
        return cell_top

    def wc(r, c, value="", **kwargs):
        cell = ws.cell(row=r, column=c, value=value)
        st(cell, **kwargs)
        return cell

    # ── LARGEURS COLONNES (exactes de l'original) ─────────
    ws.column_dimensions["A"].width  = 22.0   # MATRICULE
    ws.column_dimensions["B"].width  = 16.3   # NOMS ET PRENOMS
    ws.column_dimensions["C"].width  = 8.0    # vide
    ws.column_dimensions["D"].width  = 4.7    # jours
    for col in ["E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W"]:
        ws.column_dimensions[col].width = 4.7
    ws.column_dimensions["X"].width  = 7.7   # HN
    ws.column_dimensions["Y"].width  = 7.7   # HS
    ws.column_dimensions["Z"].width  = 8.6   # TOTAL
    ws.column_dimensions["AA"].width = 7.7   # PN
    ws.column_dimensions["AB"].width = 15.7  # OBSERVATIONS
    ws.column_dimensions["AC"].width = 19.0  # vide fin

    # ── HAUTEURS LIGNES (exactes de l'original) ───────────
    ws.row_dimensions[1].height  = 6.75
    ws.row_dimensions[2].height  = 15.25
    ws.row_dimensions[3].height  = 15.25
    ws.row_dimensions[4].height  = 15.25
    ws.row_dimensions[5].height  = 15.25
    ws.row_dimensions[6].height  = 26.25
    ws.row_dimensions[7].height  = 8.25
    ws.row_dimensions[8].height  = 15.25
    ws.row_dimensions[9].height  = 15.25
    ws.row_dimensions[10].height = 15.25
    ws.row_dimensions[11].height = 15.25
    ws.row_dimensions[12].height = 25.5
    ws.row_dimensions[13].height = 25.5
    ws.row_dimensions[14].height = 25.5

# ── LIGNE 2-8 : TITRE ─────────────────────────────────
    mc(2, 7, 8, 29,
        "POINTAGE HEBDOMADAIRE DU PERSONNEL AGENT DE MAITRISE",
        bold=True, size=14, halign="center", valign="center", border=b_thin)

    mc(2, 1, 4, 6, "USINES DE DOUALA",
        bold=True, size=11, halign="center", valign="center", border=b_thin)

    mc(6, 1, 8, 6, f"USINE/SITE : NDOKOTI",
        bold=True, size=11, halign="center", valign="center", border=b_thin)

    mc(9, 1, 11, 6, "SERVICE : CONDITIONNEMENT",
        bold=True, size=11, halign="center", valign="center", border=b_thin)

# ── LIGNE 12 : CODE + SEMAINE + DU/AU ─────────────────
    wc(12, 1, "   CODE :……", size=9, halign="left")
    wc(12, 2, f"SEMAINE {sem_num}", bold=True, size=9, halign="left")
    wc(12, 7, "Du", size=9, halign="center")
    mc(12, 10, 12, 15, date_lundi.strftime("%d/%m/%Y"), bold=True, size=9, halign="center", border=b_thick)
    wc(12, 19, "Au", size=9, halign="center")
    mc(12, 22, 12, 25, date_dimanche.strftime("%d/%m/%Y"), bold=True, size=9, halign="center", border=b_thick)

    # ── LIGNES 13-14 : EN-TÊTES TABLEAU ───────────────────
    # Colonnes exactes : D=4, G=7, J=10, M=13, P=16, S=19, V=22, Dim=22-23
    jours_cols = [
        ("Lun", 4,  6,  date_lundi + timedelta(days=0)),
        ("Mar", 7,  9,  date_lundi + timedelta(days=1)),
        ("Mer", 10, 12, date_lundi + timedelta(days=2)),
        ("Jeu", 13, 15, date_lundi + timedelta(days=3)),
        ("Ven", 16, 18, date_lundi + timedelta(days=4)),
        ("Sam", 19, 21, date_lundi + timedelta(days=5)),
        ("Dim", 22, 23, date_lundi + timedelta(days=6)),
    ]

    mc(13, 1,  14, 1,  "MATRICULE",       bold=True, size=9, halign="center", border=b_thick)
    mc(13, 2,  14, 3,  "NOMS ET PRENOMS", bold=True, size=9, halign="center", border=b_thick, wrap=True)

    for jour, c1, c2, d in jours_cols:
        mc(13, c1, 13, c2, jour, bold=True, size=9, halign="center",
           border=Border(top=thick, bottom=dashed, left=thick, right=thick))
        if jour == "Dim":
            wc(14, c1, "F",  bold=True, size=9, halign="center",
               border=Border(top=dashed, bottom=thick, left=thick, right=dashed))
            wc(14, c2, "PN", bold=True, size=9, halign="center",
               border=Border(top=dashed, bottom=thick, left=dashed, right=thick))
        else:
            wc(14, c1,   "N",  bold=True, size=9, halign="center",
               border=Border(top=dashed, bottom=thick, left=thick,  right=dashed))
            wc(14, c1+1, "F",  bold=True, size=9, halign="center",
               border=Border(top=dashed, bottom=thick, left=dashed, right=dashed))
            wc(14, c2,   "PN", bold=True, size=9, halign="center",
               border=Border(top=dashed, bottom=thick, left=dashed, right=thick))

    mc(13, 24, 14, 24, "HN",    bold=True, size=9, halign="center", border=b_thick)
    mc(13, 25, 14, 25, "HS",    bold=True, size=9, halign="center", border=b_thick)
    mc(13, 26, 14, 26, "TOTAL", bold=True, size=9, halign="center", border=b_thick)
    mc(13, 27, 14, 27, "PN",    bold=True, size=9, halign="center", border=b_thick)
    mc(13, 28, 14, 29, "OBSERVATIONS", bold=True, size=9, halign="center", border=b_thick, wrap=True)

    # ── LIGNES MEMBRES ────────────────────────────────────
    ROW_S = 15
    nb_am = max(len(membres_am), 1)
    ROW_E = ROW_S + nb_am - 1

    for idx in range(nb_am):
        r = ROW_S + idx
        ws.row_dimensions[r].height = 36.0

        if idx < len(membres_am):
            m = membres_am[idx]
            presences = {}
            for j_idx, (_, c1, c2, d) in enumerate(jours_cols):
                d_key = d.date() if hasattr(d, 'date') else d
                ptg = par_date.get(d_key)
                if ptg:
                    for lg in ptg.lignes:
                        if lg.membre_id == m.id:
                            presences[j_idx] = lg
                            break

            wc(r, 1, m.matricule or "", size=9, halign="center", border=b_thin)
            mc(r, 2, r, 3, m.nom_prenom, size=9, halign="left", border=b_thin)

            for j_idx, (jour, c1, c2, d) in enumerate(jours_cols):
                lg = presences.get(j_idx)
                if jour == "Dim":
                    f_val  = lg.heures_F  if lg and lg.heures_F  else ""
                    pn_val = lg.heures_PN if lg and lg.heures_PN else ""
                    wc(r, c1, f_val,  size=9, halign="center", border=b_thin)
                    wc(r, c2, pn_val, size=9, halign="center", border=b_thin)
                else:
                    if lg:
                        n_val  = lg.presence if lg.presence in ("AA","AB","R","RM","CP","CD","CM","CN") \
                                 else (lg.heures_N if lg.heures_N else "")
                        f_val  = lg.heures_F  if lg.heures_F  else ""
                        pn_val = lg.heures_PN if lg.heures_PN else ""
                    else:
                        n_val = f_val = pn_val = ""
                    wc(r, c1,   n_val,  size=9, halign="center", border=b_thin)
                    wc(r, c1+1, f_val,  size=9, halign="center", border=b_thin)
                    wc(r, c2,   pn_val, size=9, halign="center", border=b_thin)
        else:
            wc(r, 1, "", border=b_thin)
            mc(r, 2, r, 3, "", border=b_thin)
            for _, c1, c2, _ in jours_cols:
                for ci in range(c1, c2+1):
                    wc(r, ci, "", border=b_thin)

        # Formules identiques à l'original
        rs = str(r)
        ws.cell(r, 24).value = (
            f"=IF(SUM(D{rs},E{rs},G{rs},H{rs},J{rs},K{rs},M{rs},N{rs},"
            f"P{rs},Q{rs},S{rs},T{rs})>50,50,"
            f" SUM(D{rs},E{rs},G{rs},H{rs},J{rs},K{rs},M{rs},N{rs},"
            f"P{rs},Q{rs},S{rs},T{rs}))"
        )
        ws.cell(r, 25).value = (
            f"=SUM(D{rs},E{rs},G{rs},H{rs},J{rs},K{rs},M{rs},N{rs},"
            f"P{rs},Q{rs})+SUM(S{rs},T{rs},V{rs})-X{rs}"
        )
        ws.cell(r, 26).value = f"=SUM(X{rs}:Y{rs})"
        ws.cell(r, 27).value = f"=SUM(F{rs},I{rs},L{rs},O{rs},R{rs},U{rs},W{rs})"
        for ci in [24, 25, 26, 27]:
            st(ws.cell(r, ci), size=9, halign="center", border=b_thin)
        mc(r, 28, r, 29, "", border=b_thin)

    # ── TOTAL ─────────────────────────────────────────────
    R_TOT = ROW_E + 1
    ws.row_dimensions[R_TOT].height = 22.5
    mc(R_TOT, 19, R_TOT, 22, "TOTAL", bold=True, size=9, halign="right", border=b_thick)
    for ci, col in [(24,"X"),(25,"Y"),(26,"Z")]:
        ws.cell(R_TOT, ci).value = f"=SUM({col}{ROW_S}:{col}{ROW_E})"
        st(ws.cell(R_TOT, ci), bold=True, size=9, halign="center", border=b_thick)

    # ── TAUX H.S ──────────────────────────────────────────
    R_TAUX = R_TOT + 1
    ws.row_dimensions[R_TAUX].height = 22.5
    wc(R_TAUX, 4, f"SEM-{sem_num:02d}", bold=True, size=9, halign="left")
    mc(R_TAUX, 19, R_TAUX, 21, "TAUX H.S", bold=True, size=9, halign="right")
    ws.cell(R_TAUX, 24).value = f"=(Y{R_TOT})/Z{R_TOT}"
    ws.cell(R_TAUX, 24).number_format = "0%"
    st(ws.cell(R_TAUX, 24), bold=True, size=9, halign="center")

    # ── LÉGENDE ───────────────────────────────────────────
    R_LEG = R_TAUX + 2
    ws.row_dimensions[R_LEG].height   = 22.5
    ws.row_dimensions[R_LEG+1].height = 22.5
    ws.row_dimensions[R_LEG+2].height = 22.5

    wc(R_LEG,   1,  "N=Normal",               size=8, halign="left")
    wc(R_LEG,   4,  "PN=",                    size=8)
    wc(R_LEG,   7,  "Panier de nuit",          size=8, halign="left")
    wc(R_LEG,   16, "F=",                      size=8)
    wc(R_LEG,   19, "FERIEE",                  size=8, halign="left")

    mc(R_LEG+1, 1,  R_LEG+1, 3,  "   A=Absence injustifié", size=8, halign="left")
    wc(R_LEG+1, 4,  "CD=",                                   size=8)
    mc(R_LEG+1, 7,  R_LEG+1, 14, "         Congé décès",    size=8, halign="left")
    wc(R_LEG+1, 16, "CN=",                                   size=8)
    mc(R_LEG+1, 19, R_LEG+1, 27, "             Congé naissance", size=8, halign="left")
    wc(R_LEG+1, 28, "R=Repos",                               size=8)

    mc(R_LEG+2, 1,  R_LEG+2, 3,  "   AA=Absence autorisé",  size=8, halign="left")
    wc(R_LEG+2, 4,  "CM=",                                   size=8)
    mc(R_LEG+2, 7,  R_LEG+2, 14, "          Congé mariage", size=8, halign="left")
    wc(R_LEG+2, 16, "CP=",                                   size=8)
    mc(R_LEG+2, 19, R_LEG+2, 27, "      Congé payé",         size=8, halign="left")
    wc(R_LEG+2, 28, "RM=Maladie",                             size=8)

    # ── VISAS ─────────────────────────────────────────────
    R_VIS = R_LEG + 4
    ws.row_dimensions[R_VIS].height = 23.25
    mc(R_VIS, 1, R_VIS, 10, "Observations :", bold=True, size=9, halign="center", border=b_thick)
    mc(R_VIS, 11, R_VIS, 21, "Visa AM",              bold=True, size=8, halign="center", border=b_thick)
    mc(R_VIS, 22, R_VIS, 27, "Visa Chef d'Atelier",  bold=True, size=8, halign="center", border=b_thick)
    mc(R_VIS, 28, R_VIS, 29, "Visa Chef de Service", bold=True, size=8, halign="center", border=b_thick)

    taux_items = [
        ("Taux d'heures prévu :",   "%"),
        ("Taux d'heures realisé :", "%"),
        ("Ecart :             ",    "%"),
        ("Correspondant à : ",      "H"),
        ("Justifiées par :",        ""),
    ]
    for i, (label, unite) in enumerate(taux_items):
        r = R_VIS + 1 + i
        ws.row_dimensions[r].height = 24.0
        wc(r, 1, label, size=9, halign="left", border=b_thick)
        if unite:
            wc(r, 7, unite, size=9, border=b_thick)
    # Bordures sur toute la ligne gauche
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = b_thick

    mc(R_VIS+1, 11, R_VIS+5, 21, "", border=b_thick)
    mc(R_VIS+1, 22, R_VIS+5, 27, "", border=b_thick)
    mc(R_VIS+1, 28, R_VIS+5, 29, "", border=b_thick)

    # ── MISE EN PAGE ──────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (f"Pointage_AM_{equipe.nom}_{equipe.chaine}_"
                f"SEM{sem_num:02d}_{date_lundi.strftime('%Y-%m-%d')}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export-excel-pepiniere")
def export_excel_pepiniere(
    equipe_id:  int  = Query(...),
    date_lundi: date = Query(...),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta
    import io

    date_dimanche = date_lundi + timedelta(days=6)
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Équipe non trouvée")

    pointages = db.query(Pointage).filter(
        Pointage.equipe_id == equipe_id,
        Pointage.date      >= date_lundi,
        Pointage.date      <= date_dimanche,
    ).all()
    par_date = {p.date: p for p in pointages}

    membres_pep = sorted(
        [m for m in equipe.membres if m.statut == 'pepiniere' and m.actif != False],
        key=lambda m: m.ordre
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    sem_num = date_lundi.isocalendar()[1]
    ws.title = f"SEM{sem_num:02d}"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = 8  # A3

    # ── STYLES ──────────────────────────────────────────
    thin  = Side(style="thin",   color="000000")
    thick = Side(style="medium", color="000000")
    b_tk  = Border(top=thick, bottom=thick, left=thick, right=thick)
    b_tn  = Border(top=thin,  bottom=thin,  left=thin,  right=thin)

    def st(cell, bold=False, size=9, halign="center", valign="center",
           border=None, wrap=False):
        cell.font      = Font(bold=bold, size=size, name="Arial")
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if border: cell.border = border

    def mc(r1, c1, r2, c2, value="", **kwargs):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        st(cell, **kwargs)
        return cell

    def wc(r, c, value="", **kwargs):
        cell = ws.cell(row=r, column=c, value=value)
        st(cell, **kwargs)
        return cell

    # ── LARGEURS COLONNES (exactes depuis référence) ─────
    widths = {
        'A':1.85,'B':3.28,'C':5.42,'D':20.28,
        'E':6.71,'F':4.71,'G':6.71,'H':4.71,'I':6.71,'J':5.14,
        'K':6.71,'L':4.71,'M':6.71,'N':4.71,'O':6.71,'P':4.71,
        'Q':6.71,'R':4.71,'S':1.14,'T':6.0,'U':5.14,'V':10.28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── HAUTEURS LIGNES ──────────────────────────────────
    ws.row_dimensions[1].height = 13.5
    for r in range(2, 9):  ws.row_dimensions[r].height = 15.0
    for r in range(9, 25): ws.row_dimensions[r].height = 15.95
    for r in range(25,32): ws.row_dimensions[r].height = 15.0

    # ── LIGNE 2: TITRE ───────────────────────────────────
    mc(2,5,2,22, 'POINTAGE  OPERATEURS MACHINES AVERO USINE NDOKOTI ',
       bold=True, size=11, border=b_tk)

    # ── LIGNE 3: SERVICE ─────────────────────────────────
    mc(3,9,3,18, 'EMBOUTEILLAGE', bold=True, size=10, border=b_tn)

    # ── LIGNE 4: Chaine / Equipe / Date ─────────────────
    chain_nom = equipe.chaine if equipe.chaine else ''
    mc(4,8,4,9,  'Chaine:',  bold=True, size=9, border=b_tn)
    wc(4,10,     chain_nom,   bold=True, size=9, border=b_tn)
    mc(4,13,4,14,'Equipe:',      bold=True, size=9, border=b_tn)
    mc(4,15,4,20, equipe.nom,    bold=True, size=9, border=b_tn)
    mc(4,21,4,22, date_lundi,    bold=True, size=9, border=b_tn)

    # ── LIGNE 5: vide ────────────────────────────────────
    ws.merge_cells('F5:H5')

    # ── LIGNE 6: En-têtes jours ──────────────────────────
    JOURS_FR = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi','Dimanche']
    H_COLS   = [5,7,9,11,13,15,17]  # E, G, I, K, M, O, Q

    mc(6,1,6,2, 'Semaine:', bold=True, size=9, border=b_tn)
    wc(6,3, sem_num,         bold=True, size=9, border=b_tn)
    for jour, h_col in zip(JOURS_FR, H_COLS):
        mc(6, h_col, 6, h_col+1, jour, bold=True, size=9, border=b_tn)
    mc(6,20,6,21, 'Repartition',                   bold=True, size=9, border=b_tn)
    mc(6,22,7,22, 'Coût                       (O/N)',
       bold=True, size=8, border=b_tn, wrap=True)  # V6:V7 mergé

    # ── LIGNE 7: Dates + sous-titres répartition ─────────
    jours = [date_lundi + timedelta(days=i) for i in range(7)]

    # Lundi : date réelle, Mardi-Dim : formule =E7+1 etc.
    ws.cell(row=7, column=5).value = jours[0]
    st(ws.cell(row=7, column=5), size=9, border=b_tn)
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=6)

    prev_col_ltr = 'E'
    for i in range(1, 7):
        col = H_COLS[i]
        col_ltr = get_column_letter(col)
        ws.cell(row=7, column=col).value = f'={prev_col_ltr}7+1'
        st(ws.cell(row=7, column=col), size=9, border=b_tn)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
        prev_col_ltr = col_ltr

    mc(7,20,8,20, 'Heure  Total',  bold=True, size=8, border=b_tn, wrap=True)
    mc(7,21,8,21, 'Nbre de  jour', bold=True, size=8, border=b_tn, wrap=True)

    # ── LIGNE 8: En-têtes colonnes ───────────────────────
    wc(8,1, 'St',            bold=True, size=9, border=b_tn)
    wc(8,2, '',              size=9,             border=b_tn)
    wc(8,3, 'Poste ',        bold=True, size=9, border=b_tn)
    wc(8,4, 'Nom et Prenom', bold=True, size=9, border=b_tn)
    for h_col in H_COLS:
        wc(8, h_col,   'Heure', bold=True, size=9, border=b_tn)
        wc(8, h_col+1, 'N J',   bold=True, size=9, border=b_tn)
    wc(8,22, '', size=9, border=b_tn)  # V8 = flag coût O/N (rempli manuellement)

    # ── LIGNES MEMBRES ───────────────────────────────────
    GI_START,  GI_END  = 9,  16   # 8 slots groupe I
    GII_START, GII_END = 18, 21   # 4 slots groupe II

    # Merges col A (cosmétique comme référence)
    ws.merge_cells(start_row=9,  end_row=10, start_column=1, end_column=1)
    ws.merge_cells(start_row=11, end_row=16, start_column=1, end_column=1)
    ws.merge_cells(start_row=18, end_row=21, start_column=1, end_column=1)

    def get_heure(ptg_obj, membre_id):
        if not ptg_obj: return None
        ligne = next((l for l in ptg_obj.lignes if l.membre_id == membre_id), None)
        if not ligne: return None
        if ligne.presence == 'P':
            return ligne.heures_N if ligne.heures_N else None
        return ligne.presence  # "AA", "AB", "R", "RM", etc.

    def fill_row(r, membre):
        taux = getattr(membre, 'salaire_horaire', None) or 4600
        wc(r,2, taux,            size=9, border=b_tn)
        wc(r,3, membre.fonction or '', size=9, halign='left', border=b_tn)
        wc(r,4, membre.nom_prenom,    size=9, halign='left', border=b_tn)
        for h_col, jour_date in zip(H_COLS, jours):
            ptg  = par_date.get(jour_date)
            hval = get_heure(ptg, membre.id)
            ltr  = get_column_letter(h_col)
            wc(r, h_col,   hval,             size=9, border=b_tn)
            wc(r, h_col+1, f'=N({ltr}{r})/8', size=9, border=b_tn)
        hcols_rev = '+'.join([f'N({get_column_letter(c)}{r})' for c in reversed(H_COLS)])
        wc(r,20, f'={hcols_rev}',                      size=9, border=b_tn)
        wc(r,21, f'=ROUND(T{r}/8,1)',                  size=9, border=b_tn)
        wc(r,22, f'=IF($V$8="O",U{r}*B{r},"")',        size=9, border=b_tn)

    def fill_empty_row(r):
        wc(r,2, 4600, size=9, border=b_tn)
        for h_col in H_COLS:
            ltr = get_column_letter(h_col)
            wc(r, h_col,   None,             border=b_tn)
            wc(r, h_col+1, f'=N({ltr}{r})/8', size=9, border=b_tn)
        hcols_rev = '+'.join([f'N({get_column_letter(c)}{r})' for c in reversed(H_COLS)])
        wc(r,20, f'={hcols_rev}',     size=9, border=b_tn)
        wc(r,21, f'=ROUND(T{r}/8,1)', size=9, border=b_tn)
        wc(r,22, f'=U{r}*B{r}',       size=9, border=b_tn)

    # Remplissage Groupe I
    gi_mbrs = membres_pep[:8]
    for i, r in enumerate(range(GI_START, GI_END+1)):
        if i < len(gi_mbrs): fill_row(r, gi_mbrs[i])
        else:                 fill_empty_row(r)

    # ── EFFECTIF JOURNALIER I (ligne 17) ─────────────────
    R17 = 17
    mc(R17,1,R17,4, 'EFFECTIF  JOURNALIER I', bold=True, size=9, border=b_tk)
    for h_col in H_COLS:
        ltr = get_column_letter(h_col)
        mc(R17, h_col, R17, h_col+1,
           f'=COUNT({ltr}{GI_START}:{ltr}{GI_END})',
           bold=True, size=9, border=b_tk)
    wc(R17,20, f'=SUM(T{GI_START}:T{GI_END})', bold=True, size=9, border=b_tk)
    wc(R17,21, f'=SUM(U{GI_START}:U{GI_END})', bold=True, size=9, border=b_tk)
    wc(R17,22, f'=SUM(V{GI_START}:V{GI_END})', bold=True, size=9, border=b_tk)

    # Remplissage Groupe II
    gii_mbrs = membres_pep[8:12]
    for i, r in enumerate(range(GII_START, GII_END+1)):
        if i < len(gii_mbrs): fill_row(r, gii_mbrs[i])
        else:                  fill_empty_row(r)

    # ── EFFECTIF JOURNALIER II (ligne 22) ────────────────
    R22 = 22
    mc(R22,1,R22,4, 'EFFECTIF JOURNALIER  II', bold=True, size=9, border=b_tk)
    for h_col in H_COLS:
        ltr = get_column_letter(h_col)
        mc(R22, h_col, R22, h_col+1,
           f'=COUNT({ltr}{GII_START}:{ltr}{GII_END})',
           bold=True, size=9, border=b_tk)
    wc(R22,20, f'=SUM(T{GII_START}:T{GII_END})', bold=True, size=9, border=b_tk)
    wc(R22,21, f'=SUM(U{GII_START}:U{GII_END})', bold=True, size=9, border=b_tk)
    wc(R22,22, f'=SUM(V{GII_START}:V{GII_END})', bold=True, size=9, border=b_tk)

    # ── EFFECTIF TOTAL I+II (ligne 23) ───────────────────
    R23 = 23
    mc(R23,1,R23,4, 'EFFECTIF TOTAL JOURNALIER (I+II) ', bold=True, size=9, border=b_tk)
    for h_col in H_COLS:
        ltr = get_column_letter(h_col)
        mc(R23, h_col, R23, h_col+1,
           f'={ltr}{R17}+{ltr}{R22}',
           bold=True, size=9, border=b_tk)
    wc(R23,20, f'=T{R22}+T{R17}', bold=True, size=9, border=b_tk)
    wc(R23,21, f'=U{R22}+U{R17}', bold=True, size=9, border=b_tk)
    wc(R23,22, f'=V{R22}+V{R17}', bold=True, size=9, border=b_tk)

    # ── CIBLE (ligne 24) ─────────────────────────────────
    R24 = 24
    mc(R24,2,R24,3, 'Cible', size=9, border=b_tn)
    wc(R24,4, 26, size=9, border=b_tn)  # D24 = cible
    for h_col in H_COLS:
        ltr = get_column_letter(h_col)
        mc(R24, h_col, R24, h_col+1,
           f'=IF({ltr}{R23}>$D$24,"Mauvais","RAS")',
           size=9, border=b_tn)
    wc(R24,21, f'=ROUND(T{R24}/8,1)', size=9, border=b_tn)

    # ── LIGNE 26: Légende ────────────────────────────────
    wc(26,1, 'RM: Repos Médical', size=8)
    wc(26,4, 'A: Absence',        size=8)
    wc(26,5, 'C: Congé ',         size=8)

    # ── LIGNES 27-30: Observations + Visa ────────────────
    wc(27,1, 'Observat°', bold=True, size=9)
    mc(27, 5, 30,17, '', border=b_tn)           # zone observations E27:Q30
    mc(27,18,27,20, 'Adj. CSC',        bold=True, size=8, border=b_tn)
    mc(27,21,27,22, 'Chef de Service', bold=True, size=8, border=b_tn)
    mc(28,18,30,20, '', border=b_tn)            # zone visa Adj CSC
    mc(28,21,30,22, '', border=b_tn)            # zone visa Chef de Service

    # ── LIGNE 31: Référence ──────────────────────────────
    wc(31,21, 'BNT',     size=9)
    wc(31,22, 'Page1/6', size=9)

    # ── EXPORT ───────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Pointage_Pepiniere_{date_lundi}.xlsx"'}
    )

@router.get("/export-excel-occasionnel")
def export_excel_occasionnel(
    equipe_id:  int  = Query(...),
    date_lundi: date = Query(...),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import timedelta
    import io

    date_dimanche = date_lundi + timedelta(days=6)
    equipe = db.query(Equipe).filter(Equipe.id == equipe_id).first()
    if not equipe:
        raise HTTPException(status_code=404, detail="Équipe non trouvée")

    pointages = db.query(Pointage).filter(
        Pointage.equipe_id == equipe_id,
        Pointage.date      >= date_lundi,
        Pointage.date      <= date_dimanche,
    ).all()
    par_date = {p.date: p for p in pointages}

    membres_occ = sorted(
        [m for m in equipe.membres if m.statut == 'occasionnel' and m.actif != False],
        key=lambda m: m.ordre
    )

    if not membres_occ:
        raise HTTPException(status_code=404, detail="Aucun occasionnel dans cette équipe")

    # Semaine du mois (1 à 6) → détermine la ligne à remplir
    week_in_month = min((date_lundi.day - 1) // 7 + 1, 6)
    row_hn  = 8 + (week_in_month - 1) * 2 + 2   # HEURE NORMALE
    row_hs  = row_hn + 1                          # HEURE SUPPLEMENTAIRE
    sem_label = f"{date_lundi.strftime('%d/%m')} - {date_dimanche.strftime('%d/%m/%Y')}"

    jours = [date_lundi + timedelta(days=i) for i in range(7)]

    # ── STYLES ────────────────────────────────────────────
    thin  = Side(style="thin",   color="000000")
    thick = Side(style="medium", color="000000")
    b_tk  = Border(top=thick, bottom=thick, left=thick, right=thick)
    b_tn  = Border(top=thin,  bottom=thin,  left=thin,  right=thin)

    def st(cell, bold=False, size=10, halign="center", valign="center",
           border=None, wrap=False):
        cell.font      = Font(bold=bold, size=size, name="Arial")
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if border: cell.border = border

    def mc(ws, r1, c1, r2, c2, value="", **kwargs):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=value)
        st(cell, **kwargs)
        for row in range(r1, r2+1):
            for col in range(c1, c2+1):
                st(ws.cell(row=row, column=col), **kwargs)
        return cell

    def wc(ws, r, c, value="", **kwargs):
        cell = ws.cell(row=r, column=c, value=value)
        st(cell, **kwargs)
        return cell

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # on crée les sheets manuellement

    for membre in membres_occ:
        # Nom court pour le titre de l'onglet
        nom_court = membre.nom_prenom.split()[0][:20] if membre.nom_prenom else f"OCC{membre.id}"
        ws = wb.create_sheet(title=nom_court)

        # ── LARGEURS COLONNES (depuis référence) ─────────
        ws.column_dimensions['A'].width = 31.14
        ws.column_dimensions['B'].width = 20.57
        for c in ['C','D','E','F','G','H','I']:
            ws.column_dimensions[c].width = 8.0
        ws.column_dimensions['J'].width = 9.71
        ws.column_dimensions['K'].width = 12.71

        # ── HAUTEURS LIGNES ───────────────────────────────
        ws.row_dimensions[1].height = 16.5
        ws.row_dimensions[2].height = 15.75
        for r in [4,5,6,7]: ws.row_dimensions[r].height = 22.5
        ws.row_dimensions[9].height = 25.5
        for r in range(10, 23): ws.row_dimensions[r].height = 15.0
        ws.row_dimensions[25].height = 18.0
        ws.row_dimensions[26].height = 23.45
        ws.row_dimensions[27].height = 40.5

        # ── LIGNE 1 : Annexe ─────────────────────────────
        wc(ws, 1, 11, 'Annexe 4', size=9, halign='right')

        # ── LIGNE 2 : Titre ──────────────────────────────
        mc(ws, 2,2, 2,11,
           'FEUILLE DE POINTAGE DU TRAVAILLEUR TEMPORAIRE',
           bold=True, size=11, border=b_tk)

        # ── LIGNES 4-7 : Infos travailleur ───────────────
        infos = [
            (4, 'NOM :', membre.nom_prenom or '',         'Date & Lieu de Naissance :', ''),
            (5, 'PRENOM :', '',                            'N° CNPS :', ''),
            (6, 'MATRICULE :', membre.matricule or '',     'Catégorie Professionnelle :', membre.fonction or ''),
            (7, 'SERVICE :', equipe.nom,                   'Salaire Horaire / Mensuel :', str(membre.salaire_horaire or '')),
        ]
        for r, lbl1, val1, lbl2, val2 in infos:
            wc(ws, r, 2, lbl1, bold=True, size=9, halign='left', border=b_tn)
            mc(ws, r,3, r,5, val1, size=9, halign='left', border=b_tn)
            mc(ws, r,6, r,7, lbl2, bold=True, size=9, halign='left', border=b_tn)
            mc(ws, r,9, r,10, val2, size=9, halign='left', border=b_tn)

        # ── LIGNE 9 : En-têtes tableau ───────────────────
        headers = ['DECOMPTE DES HEURES','SEMAINE','LUNDI','MARDI',
                   'MERCREDI','JEUDI','VENDREDI','SAMEDI','DIMANCHE','TOTAL','VISA CHEF ATELIER']
        for i, h in enumerate(headers):
            wc(ws, 9, i+1, h, bold=True, size=9, border=b_tk, wrap=True)

        # ── LIGNES 10-21 : 6 semaines (HN + HS) ─────────
        for w in range(6):
            r_hn = 10 + w * 2
            r_hs = r_hn + 1
            wc(ws, r_hn, 1, 'HEURE NORMALE',        bold=True, size=9, border=b_tn)
            wc(ws, r_hs, 1, 'HEURE SUPPLEMENTAIRE', bold=True, size=9, border=b_tn)
            for r in [r_hn, r_hs]:
                # Col B : semaine (vide sauf la semaine courante)
                wc(ws, r, 2, '', border=b_tn)
                # Cols C-I : jours
                for c in range(3, 10):
                    wc(ws, r, c, None, border=b_tn)
                # Col J : TOTAL
                col_start = get_column_letter(3)
                col_end   = get_column_letter(9)
                ws.cell(row=r, column=10).value = f'=SUM({col_start}{r}:{col_end}{r})'
                st(ws.cell(row=r, column=10), size=9, border=b_tn)
                # Col K : VISA
                wc(ws, r, 11, '', border=b_tn)

        # ── REMPLIR LA SEMAINE COURANTE ───────────────────
        wc(ws, row_hn, 2, sem_label, size=8, border=b_tn)
        wc(ws, row_hs, 2, sem_label, size=8, border=b_tn)

        for d_idx, jour_date in enumerate(jours):
            col = 3 + d_idx  # C=3 (lundi) → I=9 (dimanche)
            ptg = par_date.get(jour_date)
            hn_val = hs_val = None

            if ptg:
                ligne = next(
                    (l for l in ptg.lignes if l.membre_id == membre.id), None
                )
                if ligne:
                    if ligne.presence in ('AA','AB','R','RM','CP','A','C'):
                        hn_val = ligne.presence
                    else:
                        hn_val = ligne.heures_N  if ligne.heures_N  else None
                        hs_val = (ligne.heures_F or 0) + (ligne.heures_PN or 0) or None

            wc(ws, row_hn, col, hn_val, size=9, border=b_tn)
            wc(ws, row_hs, col, hs_val, size=9, border=b_tn)

        # ── LIGNE 23 : Acompte ────────────────────────────
        mc(ws, 23,1, 23,11,
           'ACOMPTE SUR SALAIRE : _______________ Payé le : ______________',
           size=9, halign='left', border=b_tn)

        # ── LIGNES 25-27 : Signatures ─────────────────────
        mc(ws, 25,1, 27,1, 'SIGNATURE DU TRAVAILLEUR',
           bold=True, size=9, border=b_tk)
        mc(ws, 25,6, 25,9, 'CHEF DE SERVICE',
           bold=True, size=9, border=b_tk)
        wc(ws, 26,6, 'NOM :',       size=9, halign='left', border=b_tn)
        mc(ws, 26,7, 26,9, '', border=b_tn)
        wc(ws, 27,6, 'SIGNATURE :', size=9, halign='left', border=b_tn)
        mc(ws, 27,7, 27,9, '', border=b_tn)

        # Mise en page
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize   = 9   # A4

    # ── EXPORT ───────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    sem_num = date_lundi.isocalendar()[1]
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition':
                 f'attachment; filename="Pointage_Occasionnel_SEM{sem_num}_{date_lundi}.xlsx"'}
    )

@router.get("/rapport-ab")
def rapport_absences_non_justifiees(
    date_debut: date = Query(...),
    date_fin:   date = Query(...),
    chaine:     Optional[str] = Query(None),
    equipe_id:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user)
):
    """Rapport des absences non justifiées (AB) sur une période"""
    query = db.query(Pointage).filter(
        Pointage.date >= date_debut,
        Pointage.date <= date_fin,
    )
    if chaine:    query = query.filter(Pointage.chaine    == chaine)
    if equipe_id: query = query.filter(Pointage.equipe_id == equipe_id)
    
    pointages = query.all()
    
    # Regrouper les AB par membre
    rapport = {}
    for ptg in pointages:
        for lg in ptg.lignes:
            if lg.presence == "AB" and not lg.est_occasionnel:
                cle = f"{lg.nom_prenom}_{lg.membre_id}"
                if cle not in rapport:
                    rapport[cle] = {
                        "nom_prenom":   lg.nom_prenom,
                        "membre_id":    lg.membre_id,
                        "fonction":     lg.fonction,
                        "chaine":       ptg.chaine,
                        "equipe_nom":   ptg.equipe_nom,
                        "nb_absences":  0,
                        "heures_total": 0,
                        "details":      [],
                    }
                rapport[cle]["nb_absences"]  += 1
                rapport[cle]["heures_total"] += ptg.quart and {
                    "7h-19h": 12, "19h-7h": 12,
                    "7h-14h": 7,  "14h-21h": 7, "21h-7h": 10,
                }.get(ptg.quart, 8) or 8
                rapport[cle]["details"].append({
                    "date":  str(ptg.date),
                    "quart": ptg.quart,
                    "jour":  ptg.date.strftime("%A %d/%m/%Y"),
                })
    
    # Trier par nb absences décroissant
    result = sorted(rapport.values(), key=lambda x: x["nb_absences"], reverse=True)
    return result